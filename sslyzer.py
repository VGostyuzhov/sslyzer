#!/usr/bin/python2
# -*- coding: utf-8 -*-
#Import standard libraries
import os
import sys
import argparse
import json
from pprint import pprint

#Import sslyze Classes
from sslyze.server_connectivity import ServerConnectivityInfo, ServerConnectivityError
from sslyze.synchronous_scanner import SynchronousScanner
from sslyze.concurrent_scanner import ConcurrentScanner, PluginRaisedExceptionScanResult
from sslyze.plugins.openssl_cipher_suites_plugin import Sslv20ScanCommand, Sslv30ScanCommand, Tlsv10ScanCommand, Tlsv11ScanCommand, Tlsv12ScanCommand
from sslyze.plugins.heartbleed_plugin import HeartbleedScanCommand
from sslyze.plugins.compression_plugin import CompressionScanCommand
from sslyze.plugins.fallback_scsv_plugin import FallbackScsvScanCommand
from sslyze.plugins.openssl_ccs_injection_plugin import OpenSslCcsInjectionScanCommand
from sslyze.plugins.session_renegotiation_plugin import SessionRenegotiationScanCommand
from sslyze.plugins.certificate_info_plugin import CertificateInfoScanCommand
from sslyze.plugins.utils.certificate_utils import CertificateUtils

#Import Python packages
from termcolor import colored
from openpyxl import Workbook
from openpyxl.styles import NamedStyle

#Import custom modules
from styles import styles

BOLD = '\033[1m'
ENDBOLD = '\033[0m'


class ConnectionError(Exception):
    pass

def get_weak_ciphers(server):
    rc4_ciphers = []
    des_ciphers = []
    anon_ciphers = []
    weak_hash_ciphers = []
    dh_export_ciphers = []
    for protocol, ciphers in server['cipher_suites'].iteritems():
        for cipher in ciphers:
            cipher = cipher['cipher']
            if 'RC4' in cipher and cipher not in rc4_ciphers:
                rc4_ciphers.append(cipher)
            if (('MD5' or 'SHA-') in cipher or cipher.endswith('SHA')) and cipher not in weak_hash_ciphers:
                weak_hash_ciphers.append(cipher)
            if 'DES' in cipher and cipher not in des_ciphers:
                des_ciphers.append(cipher)
            if 'anon' in cipher and cipher not in anon_ciphers:
                anon_ciphers.append(cipher)
            if 'dh_info' in cipher:
                if cipher['dh_info']['Type'] == 'DH' and int(cipher['dh_info']['GroupSize']) <= 1024:
                    if cipher not in dh_export_ciphers:
                        dh_export_ciphers.append(cipher)
    return {'rc4': rc4_ciphers, 'des': des_ciphers, 'dh_export': dh_export_ciphers, 'weak_hash': weak_hash_ciphers}

def check_Beast(server):
    protocols = {key: server['cipher_suites'][key] for key in ['sslv30', 'tlsv10']}
    for protocol, ciphers in protocols.iteritems():
        for cipher in ciphers:
            if 'CBC' in cipher['cipher']:
                return True
    return False

def check_Freak(server):
    export_rsa_ciphers = ['EXP1024-DES-CBC-SHA',
            'EXP1024-RC2-CBC-MD5',
            'EXP1024-RC4-SHA',
            'EXP1024-RC4-MD5',
            'EXP-EDH-RSA-DES-CBC-SHA',
            'EXP-DH-RSA-DES-CBC-SHA',
            'EXP-DES-CBC-SHA',
            'EXP-RC2-CBC-MD5',
            'EXP-RC4-MD5']
    for protocol, ciphers in server['cipher_suites'].iteritems():
        for cipher in ciphers:
            if cipher in export_rsa_ciphers:
                return True
    return False

def scan_server(hostname):
    """Scan server for supported SSL cipher suites and vulnerabilities and
    return dict object"""
    server = {'hostname': hostname, 'cipher_suites': {}, 'weak_ciphers': {}, 'cert': {}}
    vulners = {'heartbleed': None, 'crime': None, 'downgrade': None}
    # Setup the server to scan and ensure it is online/reachable
    print(colored('Processing host {0}{1}{2}'.format(BOLD, hostname, ENDBOLD), 'cyan'))
    sys.stdout.write('Testing connectivity: ')
    #Python buffer strings before printing to stdout. Let's fush it to screen before process next command.
    sys.stdout.flush()

    try:
        server_info = ServerConnectivityInfo(hostname)
        server_info.test_connectivity_to_server()
        sys.stdout.write(colored(BOLD + 'OK!\n' + ENDBOLD, 'green'))
    except ServerConnectivityError as e:
        #raise ConnectionError('Error when connecting to {}: {}'.format(hostname, e.error_msg))
        raise 

    sys.stdout.write('Getting test results: ')
    sys.stdout.flush()
    
    concurrent_scanner = ConcurrentScanner(network_retries=3, network_timeout=10)

    concurrent_scanner.queue_scan_command(server_info, Sslv20ScanCommand())
    concurrent_scanner.queue_scan_command(server_info, Sslv30ScanCommand())
    concurrent_scanner.queue_scan_command(server_info, Tlsv10ScanCommand())
    concurrent_scanner.queue_scan_command(server_info, Tlsv11ScanCommand())
    concurrent_scanner.queue_scan_command(server_info, Tlsv12ScanCommand())
    concurrent_scanner.queue_scan_command(server_info, HeartbleedScanCommand())
    concurrent_scanner.queue_scan_command(server_info, CompressionScanCommand())
    concurrent_scanner.queue_scan_command(server_info, FallbackScsvScanCommand())
    concurrent_scanner.queue_scan_command(server_info, OpenSslCcsInjectionScanCommand())
    concurrent_scanner.queue_scan_command(server_info, SessionRenegotiationScanCommand())
    concurrent_scanner.queue_scan_command(server_info, CertificateInfoScanCommand())

    for scan_result in concurrent_scanner.get_results():
        #print('\nReceived scan result for {} on host {}'.format(scan_result.scan_command.__class__.__name__, scan_result.server_info.hostname))
        if isinstance(scan_result, PluginRaisedExceptionScanResult):
            #raise RuntimeError('Scan command failed: {}'.format(scan_result.as_text()))
            return None

        commands = {'sslv20': Sslv20ScanCommand, 'sslv30': Sslv30ScanCommand, 'tlsv10': Tlsv10ScanCommand, 'tlsv11': Tlsv11ScanCommand, 'tlsv12': Tlsv12ScanCommand}
        for protocol, command in commands.iteritems():
            if isinstance(scan_result.scan_command, command):
                server['cipher_suites'][protocol] = []
                for cipher in scan_result.accepted_cipher_list:
                    c = {'cipher': cipher.name}
                    if cipher.dh_info is not None:
                        if cipher.dh_info['Type'] == 'DH':
                            c['dh_info'] = {key: cipher.dh_info[key] for key in ['Type', 'GroupSize']}
                        elif cipher.dh_info['Type'] == 'ECDH':
                            c['dh_info'] = {key: cipher.dh_info[key] for key in ['Type', 'GroupSize']}
                    server['cipher_suites'][protocol].append(c)
                server[protocol] = bool(len(server['cipher_suites'][protocol]))

        if isinstance(scan_result.scan_command, HeartbleedScanCommand):
           vulners['heartbleed'] = scan_result.is_vulnerable_to_heartbleed

        if isinstance(scan_result.scan_command, CompressionScanCommand):
            server['compression'] = scan_result.compression_name
            vulners['crime'] = bool(server['compression'])

        if isinstance(scan_result.scan_command, FallbackScsvScanCommand):
            server['fallback_scsv'] = scan_result.supports_fallback_scsv
            vulners['downgrade'] = server['fallback_scsv']

        if isinstance(scan_result.scan_command, OpenSslCcsInjectionScanCommand):
            vulners['ccs_injection'] = scan_result.is_vulnerable_to_ccs_injection

        if isinstance(scan_result.scan_command, SessionRenegotiationScanCommand):
            server['client_reneg'] = scan_result.accepts_client_renegotiation
            server['secure_reneg'] = scan_result.supports_secure_renegotiation

        if isinstance(scan_result.scan_command, CertificateInfoScanCommand):
            server['cert']['matches_hostname'] = scan_result.certificate_matches_hostname
            cert = scan_result.certificate_chain[0]
            server['cert']['not_valid_after'] = cert.not_valid_after.strftime('%d.%m.%Y')
            subject = CertificateUtils.get_name_as_short_text(cert.subject)
            issuer = CertificateUtils.get_name_as_short_text(cert.issuer)
            if issuer == subject:
                server['cert']['self_signed'] = True
            else:
                server['cert']['self_signed'] = False
            server['cert']['trusted'] = scan_result.path_validation_result_list[0].is_certificate_trusted


    server['weak_ciphers'] = get_weak_ciphers(server)

    vulners['poodle'] = server['sslv30']
    vulners['drown'] = server['sslv20']
    vulners['rc4'] = bool(len(server['weak_ciphers']['rc4']))
    vulners['logjam'] = bool(len(server['weak_ciphers']['dh_export']))
    vulners['beast'] = check_Beast(server)
    vulners['freak'] = check_Freak(server)
    server['vulners'] = vulners
    with open('log.json', 'a') as logfile:
        server_json = json.dumps({'hostname':server['hostname'], 'vulners': server['vulners']})
        logfile.write('{}\n'.format(server_json))
    return server

def make_report(servers, filename):
    try:
        wb = Workbook()
    except Exception, error:
        raise error
    for key, value in styles.iteritems():
        wb.add_named_style(value)

    sheet = wb.create_sheet('SSL')
    headers = {'Domain': 1, 'SSLv2': 2, 'SSLv3': 3, 'TLSv1.0': 4, 'TLSv1.1': 5, 'TLSv1.2': 6,
            'Heartbleed': 7, 'Crime': 8, 'Downgrade': 9, 'Poodle': 10 , 'RC4': 11, 'Beast': 12, 'CCS Injection': 13, 'Drown': 14, 'Freak': 15, 'Logjam': 16,
            'Trusted': 17, 'Self Signed': 18, 'Valid to': 19, 'Matches hostname': 20}
    for key, value in headers.iteritems():
        sheet.cell(column=value, row=1, value=key)
        sheet.cell(column=value, row=1).style = styles['Header']
    row = 2
    for server in servers:
        if hostname in server:
            sheet.cell(column=1, row=row, value=server['hostname'])
            sheet.cell(column=2, row=row, value=str(server['sslv20']))
            sheet.cell(column=3, row=row, value=str(server['sslv30']))
            sheet.cell(column=4, row=row, value=str(server['tlsv10']))
            sheet.cell(column=5, row=row, value=str(server['tlsv11']))
            sheet.cell(column=6, row=row, value=str(server['tlsv12']))
            sheet.cell(column=7, row=row, value=str(server['vulners']['heartbleed']))
            sheet.cell(column=8, row=row, value=str(server['vulners']['crime']))
            sheet.cell(column=9, row=row, value=str(server['vulners']['downgrade']))
            sheet.cell(column=10, row=row, value=str(server['vulners']['poodle']))
            sheet.cell(column=11, row=row, value=str(server['vulners']['rc4']))
            sheet.cell(column=12, row=row, value=str(server['vulners']['beast']))
            sheet.cell(column=13, row=row, value=str(server['vulners']['ccs_injection']))
            sheet.cell(column=14, row=row, value=str(server['vulners']['drown']))
            sheet.cell(column=15, row=row, value=str(server['vulners']['freak']))
            sheet.cell(column=16, row=row, value=str(server['vulners']['logjam']))
            sheet.cell(column=17, row=row, value=str(server['cert']['trusted']))
            sheet.cell(column=18, row=row, value=str(server['cert']['self_signed']))
            sheet.cell(column=19, row=row, value=str(server['cert']['not_valid_after']))
            sheet.cell(column=20, row=row, value=str(server['cert']['matches_hostname']))
            row += 1

    width_dict = {'A': 15, 'B': 7, 'C': 7, 'D': 7, 'E': 7, 'F': 7, 'G': 10}
    for key, value in width_dict.iteritems():
        sheet.column_dimensions[key].width = value

    try:
        wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
        wb.save(filename)
    except Exception, error:
        raise error

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Test multiple hosts for SSL vulnerabilities and misconfigurations using sslyze library')
    parser.add_argument('-f', '--file', dest='input_file', help='File containing input hostnames', metavar='FILENAME')
    parser.add_argument('-x', '--xlsx', dest='xlsx_file', help='Save report to .xlsx file', metavar='FILENAME')
    parser.add_argument('hostname', nargs='?', help='Single hostname to scan')
    args = parser.parse_args()

    hostnames = []
    if args.input_file:
        with open(args.input_file, 'rb') as input_file:
            hostnames = [h.strip() for h in input_file.readlines()]
    elif args.hostname:
        hostnames.append(args.hostname)
    else:
        parser.print_help()
        sys.exit(0)

    servers = []
    for hostname in hostnames:
        try:
            servers.append(scan_server(hostname))
            sys.stdout.write(colored(BOLD + 'OK!\n\n'.format(hostname) + ENDBOLD, 'green'))
        except ServerConnectivityError as e:
            sys.stdout.write(colored(BOLD + 'Error when connecting to {}: {}\n\n'.format(hostname, e.error_msg) + ENDBOLD, 'red'))
            continue
    
    if args.xlsx_file:
        make_report(servers, args.xlsx_file)
    else:
        pprint(servers)

