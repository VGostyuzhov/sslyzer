from openpyxl.styles import fills, PatternFill, Border, Side, Alignment, Font, NamedStyle
from copy import copy

styles = {}
styles['Header'] = NamedStyle(name='header')
styles['excel_Bad'] = NamedStyle(name='excel_Bad')
styles['default'] = NamedStyle(name='default')

styles['Header'].font = Font(name='Calibri', size=10, bold=True, color='00000000')
styles['Header'].fill = PatternFill(fill_type=fills.FILL_SOLID, start_color='DCE6F1')
styles['Header'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
styles['Header'].border = Border(left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000'))

styles['default'].font = Font(name='Calibri', size=10, bold=False, color='00000000')
styles['default'].alignment = Alignment(horizontal='center', vertical='center')

styles['excel_Bad'].font = Font(name='Calibri', size=10, bold=True, color='E80B0B')
styles['excel_Bad'].alignment = Alignment(horizontal='center', vertical='center')
styles['excel_Bad'].fill = PatternFill(fill_type=fills.FILL_SOLID, start_color='FFCCCC')



columnWidth = {'A': 25, 'B': 14, 'C': 7, 'D': 7, 'E': 7, 'F': 7, 'G': 10,
              'H': 9, 'I': 11, 'J': 9, 'K': 10, 'Q': 7, 'V':12, 'X': 13, 'AA': 10, 'AC': 10 
              }

class Colors():
    PURPLE = '\033[95m'
    CYAN = '\033[96m'
    DARKCYAN = '\033[36m'
    BLUE = '\033[94m'
    GREEN = '\033[92m'
    YELLOW = '\033[93m'
    RED = '\033[91m'
    UNDERLINE = '\033[4m'
    BOLD = '\033[1m'
    ENDBOLD = '\033[0m'