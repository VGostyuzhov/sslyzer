import csv
from datetime import datetime
# Third-party modules
from openpyxl import Workbook
from openpyxl.styles import NamedStyle
from termcolor import colored
from pprint import pprint
# Custom modules
from styles import styles, Colors, columnWidth


def build_report(servers):
    """Build report dict from list of servers"""
    # Load data from csv to "report" dict
    with open('data/descriptions.csv', 'rb') as descr_file:
        report = {}
        csvreader = csv.reader(descr_file, delimiter=";")
        for row in csvreader:
            vulner = row[0]
            descr = row[1]
            risk_level = row[2]
            report[vulner] = {}
            report[vulner]['descr'] = descr.replace('\\n', '\n')
            report[vulner]['risk_level'] = risk_level
            report[vulner]['hosts'] = []
    # Fill hosts for vulners
    for server in servers:
        if server.vulners['Heartbleed'] == True:
            report['Heartbleed']['hosts'].append(server.hostname)
        if server.vulners['Crime'] == True:
            report['Crime']['hosts'].append(server.hostname)
        if server.vulners['Downgrade'] == True:
            report['Downgrade']['hosts'].append(server.hostname)
        if server.vulners['Poodle'] == True:
            report['Poodle']['hosts'].append(server.hostname)
        if server.vulners['RC4'] == True:
            report['RC4']['hosts'].append(server.hostname)
        if server.vulners['Beast'] == True:
            report['Beast']['hosts'].append(server.hostname)
        if server.vulners['CCS Injection'] == True:
            report['CCS Injection']['hosts'].append(server.hostname)
        if server.vulners['Drown'] == True:
            report['Drown']['hosts'].append(server.hostname)
        if server.vulners['Freak'] == True:
            report['Freak']['hosts'].append(server.hostname)
        if server.vulners['Logjam'] == True:
            report['Logjam']['hosts'].append(server.hostname)
        if server.dh_params['Common prime'] == True:
            report['DH_common_primes']['hosts'].append(server.hostname)
        if server.dh_params['Weak'] == True:
            report['DH_weak']['hosts'].append(server.hostname)
        if server.dh_params['Insecure'] == True:
            report['DH_insecure']['hosts'].append(server.hostname)
        if server.hsts_header == True:
            report['HSTS']['hosts'].append(server.hostname)
        if server.secure_reneg == True:
            report['SecureRenegotiation']['hosts'].append(server.hostname)
        if server.weak_ciphers == True:
            report['WeakCiphers']['hosts'].append(server.hostname)
        if server.insec_ciphers == True:
            report['InsecureCiphers']['hosts'].append(server.hostname)
        if server.cert.self_signed == True:
            report['SelfSigned']['hosts'].append(server.hostname)

        valid_to = datetime.strptime(server.cert.valid_to, '%d.%m.%Y')
        today = datetime.today()
        if valid_to <= today:
            report['Expired']['hosts'].append(server.hostname)
        if server.protocol_by_name('SSLv2.0').is_supported:
            report['SSLv2.0']['hosts'].append(server.hostname)
        if server.protocol_by_name('SSLv3.0').is_supported:
            report['SSLv3.0']['hosts'].append(server.hostname)
        if server.protocol_by_name('TLSv1.0').is_supported:
            report['TLSv1.0']['hosts'].append(server.hostname)
        if not server.protocol_by_name('TLSv1.2').is_supported:
            report['TLSv1.2']['hosts'].append(server.hostname)
    return report


def word_report(wb, report):
    sheet = wb.create_sheet('Word table')
    sheet.cell(column=1, row=1, value='Description')
    sheet.cell(column=2, row=1, value='Severity')
    sheet.cell(column=3, row=1, value='Hosts')
    sheet.cell(column=1, row=1).style = styles['Header']
    sheet.cell(column=2, row=1).style = styles['Header']
    sheet.cell(column=3, row=1).style = styles['Header']
    row = 2
    for key, vulner in report.iteritems():
        if len(vulner['hosts']):
            if len(vulner['hosts']) > 7:
                hosts = '{} hosts.'.format(len(vulner['hosts']))
            else:
                hosts = '\n'.join(vulner['hosts'])
            sheet.cell(column=1, row=row, value=vulner['descr'])
            sheet.cell(column=2, row=row, value=vulner['risk_level'])
            sheet.cell(column=3, row=row, value=hosts)
            row += 1
    sheet.column_dimensions['A'].width = 50
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 30


def pretty(s):
    """ """
    if s == 'Error':
        return 'Error'
    elif s == True:
        return 'Yes'
    else:
        return 'No'


def server_console_output(server):
    for protocol in server.protocols:
        if protocol.is_supported:
            print(colored(protocol.name, 'cyan'))
            print(colored(Colors.BOLD + Colors.UNDERLINE + '{:<{}}|{:<5}|{:<5}|{:<6}|{:<6}|{:<5}|{:<5}|{:<9}|{:<9}|{:<4}|{:<13}|{:<8}|{:<9}'.format(
                'Cipher', 36, 'RC4', 'MD5', 'SHA1', 'DES', 'Anon', 'PFS', 'Key Size', 'GroupSize', 'DHE', 'Common Prime', 'DH Weak', 'DH Insec') + Colors.ENDBOLD, 'white'))
            for cipher in protocol.cipher_suites:
                print('{:<36}|{:<5}|{:<5}|{:<6}|{:<6}|{:<5}|{:<5}|{:<9}|{:<9}|{:<4}|{:<13}|{:<8}|{:<9}'.format(
                    cipher.name,
                    pretty(cipher.rc4),
                    pretty(cipher.md5),
                    pretty(cipher.sha1),
                    pretty(cipher.des),
                    pretty(cipher.anon),
                    pretty(cipher.pfs),
                    cipher.key_size,
                    cipher.dh_group_size,
                    pretty(cipher.dh_export),
                    pretty(cipher.dh_common_prime),
                    pretty(cipher.dh_weak),
                    pretty(cipher.dh_insecure) ) )
            print('-----------------------------------------------------------------------------------------------------------------------------------')

    cert = server.cert
    print(Colors.BOLD + 'Certificate' + Colors.ENDBOLD)
    print(Colors.BOLD + Colors.UNDERLINE + '{:<8}|{:<8}|{:<11}|{:<11}|{:<10}|{:<10}'.format('Matches', 'Trusted', 'Valid to', 'SelfSigned', 'Hash algo', 'Weak algo') + Colors.ENDBOLD)
    print('{:<8}|{:<8}|{:<11}|{:<11}|{:<10}|{:<10}'.format(
        pretty(cert.matches_hostname),
        pretty(cert.trusted),
        cert.valid_to,
        pretty(cert.self_signed),
        cert.sign_hash_algo.upper(),
        pretty(cert.weak_hash_algo) ))
    print

    print(Colors.BOLD + 'Vulnerabilities' + Colors.ENDBOLD)
    for vulner, state in server.vulners.iteritems():
        print('{0:<17}{1}'.format(vulner, state))

    print(Colors.BOLD + 'DH params' + Colors.ENDBOLD)
    for dh_param, state in server.dh_params.iteritems():
        print('{0:<17}{1}'.format(dh_param, state))
    print
    print('{:<17}{}'.format('Forward Secrecy', str(server.pfs)))
    print('{:<17}{}'.format('HSTS', server.hsts_header))
    print('{:<17}{}'.format('Secure Reneg', str(server.secure_reneg)))
    print
    print('{:<17}{}'.format('Weak ciphers', str(server.weak_ciphers)))
    print('{:<17}{}'.format('Insecure ciphers', str(server.insec_ciphers)))
    print


def excel_report(wb, servers):
    sheet = wb.create_sheet('SSL')
    headers = {'Server': 1, 'IP': 2, 'Port': 3,
               'SSLv2': 4, 'SSLv3': 5, 'TLSv1.0': 6, 'TLSv1.1': 7, 'TLSv1.2': 8,
               'Heartbleed': 9, 'Crime': 10, 'Downgrade': 11, 'Poodle': 12 , 'RC4': 13, 'Beast': 14, 'CCS Injection': 15, 'Drown': 16, 'Freak': 17, 'Logjam': 18,
               'DH\nCommon Primes': 19, 'DH\nWeak': 20, 'DH\nInsecure': 21,
               'Forward\nSecrecy': 22, 'HSTS': 23, 'Secure\nRenegotiation': 24,
               'Weak\nCiphers': 25, 'Insecure\nCiphers': 26,
               'Trusted': 27, 'Self\nSigned': 28, 'Valid to': 29, 'Matches\nhostname': 30}
    for key, value in headers.iteritems():
        sheet.cell(column=value, row=1, value=key)
        sheet.cell(column=value, row=1).style = styles['Header']
    row = 2
    for server in servers:
        sheet.cell(column=1, row=row, value=server.hostname)
        sheet.cell(column=2, row=row, value=server.ip)
        sheet.cell(column=3, row=row, value=server.port)
        sheet.cell(column=4, row=row, value=str(server.protocol_by_name('SSLv2.0').is_supported))
        sheet.cell(column=5, row=row, value=str(server.protocol_by_name('SSLv3.0').is_supported))
        sheet.cell(column=6, row=row, value=str(server.protocol_by_name('TLSv1.0').is_supported))
        sheet.cell(column=7, row=row, value=str(server.protocol_by_name('TLSv1.1').is_supported))
        sheet.cell(column=8, row=row, value=str(server.protocol_by_name('TLSv1.2').is_supported))
        sheet.cell(column=9, row=row, value=str(server.vulners['Heartbleed']))
        sheet.cell(column=10, row=row, value=str(server.vulners['Crime']))
        sheet.cell(column=11, row=row, value=str(server.vulners['Downgrade']))
        sheet.cell(column=12, row=row, value=str(server.vulners['Poodle']))
        sheet.cell(column=13, row=row, value=str(server.vulners['RC4']))
        sheet.cell(column=14, row=row, value=str(server.vulners['Beast']))
        sheet.cell(column=15, row=row, value=str(server.vulners['CCS Injection']))
        sheet.cell(column=16, row=row, value=str(server.vulners['Drown']))
        sheet.cell(column=17, row=row, value=str(server.vulners['Freak']))
        sheet.cell(column=18, row=row, value=str(server.vulners['Logjam']))
        sheet.cell(column=19, row=row, value=str(server.dh_params['Common prime']))
        sheet.cell(column=20, row=row, value=str(server.dh_params['Weak']))
        sheet.cell(column=21, row=row, value=str(server.dh_params['Insecure']))
        sheet.cell(column=22, row=row, value=str(server.pfs))
        sheet.cell(column=23, row=row, value=str(server.hsts_header))
        sheet.cell(column=24, row=row, value=str(server.secure_reneg))
        sheet.cell(column=25, row=row, value=str(server.weak_ciphers))
        sheet.cell(column=26, row=row, value=str(server.insec_ciphers))
        sheet.cell(column=27, row=row, value=str(server.cert.trusted))
        sheet.cell(column=28, row=row, value=str(server.cert.self_signed))
        sheet.cell(column=29, row=row, value=str(server.cert.valid_to))
        sheet.cell(column=30, row=row, value=str(server.cert.matches_hostname))

        
        for i in range(1, 31):
            sheet.cell(row=row, column=i).style = styles['default']
        bad_true_columns = [4,5,6] + range(9,22) + [25, 26] + [28]
        for i in bad_true_columns:
            if sheet.cell(row=row, column=i).value == 'True':
                sheet.cell(row=row, column=i).style = styles['excel_Bad']
        bad_false_columns = [8, 22, 23, 24, 27, 30]
        for i in bad_false_columns:
            if sheet.cell(row=row, column=i).value == 'False':
                sheet.cell(row=row, column=i).style = styles['excel_Bad']        
        if server.cert.expired:
            sheet.cell(row=row, column=29).style = styles['excel_Bad']        
        #for i in range()
        row += 1


    for key, value in columnWidth.iteritems():
        sheet.column_dimensions[key].width = value


def save_report(servers, filename, report={}):
    try:
        wb = Workbook()
    except Exception, error:
        raise error
    # Register styles in workbook
    for key, value in styles.iteritems():
        wb.add_named_style(value)
    excel_report(wb, servers)
    if report:
        word_report(wb, report)
    try:
        wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
        wb.save(filename)
    except Exception, error:
        raise error
